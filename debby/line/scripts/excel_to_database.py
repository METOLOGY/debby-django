from openpyxl import load_workbook
from line.models import QAModel

wb2 = load_workbook('../chat_table/Chatbot Database.xlsx')  # path to .xlsx
chat_table = wb2['問候']  # sheet name

for num, line  in enumerate(chat_table.iter_rows()):
    if line[0].value != None and num > 0:
        question = str(line[0].value)
        for num, phrase in enumerate(line):
            if phrase.value != None:
                if num > 0:
                    answer = phrase.value
                    QA = QAModel()
                    QA.phrase = question
                    QA.answer = answer
                    QA.callback = 'Chat'
                    QA.save()