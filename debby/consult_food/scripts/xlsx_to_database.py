import openpyxl
from consult_food.models import ConsultFoodModel
from line.models import EventModel

wb = openpyxl.load_workbook("./chat_table/FoodQuery_database.xlsx", data_only=True)
ws = wb.active
food = {
    'sample_name': '',
    'modified_calorie': '',
    'carbohydrates': '',
    'dietary_fiber': '',
    'metabolic_carbohydrates': '',
    'carbohydrates_equivalent': '',
    'white_rice_equivalent': '',
}

for i in range(2, ws.max_row + 1):
    print(i)
    food['sample_name'] = ws.cell(row=i, column=1).value
    food['modified_calorie'] = ws.cell(row=i, column=2).value
    food['carbohydrates'] = ws.cell(row=i, column=3).value
    food['dietary_fiber'] = ws.cell(row=i, column=4).value
    food['metabolic_carbohydrates'] = ws.cell(row=i, column=5).value
    food['carbohydrates_equivalent'] = ws.cell(row=i, column=6).value
    food['white_rice_equivalent'] = ws.cell(row=i, column=7).value
    food.update((k, v) for k, v in food.items() if v is not None)  # remove none key value pairs
    ConsultFoodModel.objects.get_or_create(**food)
    # Event model
    event = EventModel()
    event.phrase = food['sample_name']
    event.callback = 'ConsultFood'
    event.action = 'READ'
    event.save()
