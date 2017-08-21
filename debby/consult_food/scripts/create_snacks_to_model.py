import openpyxl
from openpyxl.worksheet import Worksheet

from consult_food.models import NutritionModel, TaiwanSnackModel, NutritionTuple


# python manage.py runscript add_snacks_to_model --traceback

def split_to_list(synonyms: str):
    synonyms = synonyms.replace(' ', '')
    return synonyms.split(',')


def calc_calories(protein: float, fat: float, carbohydrates: float):
    return protein * 4.0 + fat * 9.0 + carbohydrates * 4.0


def run():
    wb = openpyxl.load_workbook("../chat_table/0807整理-台灣小吃營養大解析.xlsx", data_only=True)
    sheet_names = wb.get_sheet_names()
    for sheet_name in sheet_names:
        ws = wb.get_sheet_by_name(sheet_name)  # type: Worksheet

        row = 2
        while ws.cell(row=row, column=2).value:
            place = ws.cell(row=row, column=1).value
            place.rstrip('小吃').rstrip('伴手禮')
            print(row)
            snack_name = ws.cell(row=row, column=2).value
            print(snack_name)
            synonyms = ws.cell(row=row, column=3).value
            synonym_list = []
            if synonyms:
                synonym_list = split_to_list(synonyms)
            gram = ws.cell(row=row, column=4).value
            protein = ws.cell(row=row, column=5).value
            fat = ws.cell(row=row, column=6).value
            carbohydrates = ws.cell(row=row, column=7).value
            vegetable_amount = ws.cell(row=row, column=8).value
            oil_amount = ws.cell(row=row, column=9).value
            protein_food_amount = ws.cell(row=row, column=10).value
            grain_amount = ws.cell(row=row, column=11).value
            count_word = ws.cell(row=row, column=14).value

            nutrition = NutritionTuple(
                name=snack_name,
                gram=gram,
                calories=calc_calories(protein, fat, carbohydrates),
                protein=protein,
                fat=fat,
                carbohydrates=carbohydrates,
                fruit_amount=0,
                vegetable_amount=vegetable_amount,
                grain_amount=grain_amount,
                protein_food_amount=protein_food_amount,
                diary_amount=0,
                oil_amount=oil_amount,
            )
            nutrition_model = NutritionModel.objects.create(**nutrition._asdict())
            taiwan_snack_model = TaiwanSnackModel.objects.create(name=snack_name,
                                                                 place=place,
                                                                 count_word=count_word,
                                                                 nutrition=nutrition_model
                                                                 )
            taiwan_snack_model.synonyms.create(synonym=snack_name)
            for synonym in synonym_list:
                taiwan_snack_model.synonyms.create(synonym=synonym)

            row += 1
