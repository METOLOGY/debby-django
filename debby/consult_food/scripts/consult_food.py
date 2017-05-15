import openpyxl
from openpyxl.worksheet import Worksheet

from consult_food.models import FoodModel, FoodNameModel
from line.models import EventModel

wb = openpyxl.load_workbook("../chat_table/consult_food_database.xlsx", data_only=True)
ws = wb.active  # type: Worksheet
food = {
    'integration_number': '',
    'food_type': '',
    'sample_name': '',
    'modified_calorie': '',
    'crude_protein': '',
    'crude_fat': '',
    'carbohydrates': '',
    'dietary_fiber': '',
    'metabolic_carbohydrates': ''
}

for i in range(2, ws.max_row + 1):
    print(ws.cell(row=i, column=1).value)
    food['integration_number'] = ws.cell(row=i, column=1).value
    food['food_type'] = ws.cell(row=i, column=2).value
    food['sample_name'] = ws.cell(row=i, column=3).value
    food['modified_calorie'] = ws.cell(row=i, column=5).value
    food['crude_protein'] = ws.cell(row=i, column=6).value
    food['crude_fat'] = ws.cell(row=i, column=7).value
    food['carbohydrates'] = ws.cell(row=i, column=8).value
    food['dietary_fiber'] = ws.cell(row=i, column=9).value
    food['metabolic_carbohydrates'] = ws.cell(row=i, column=10).value

    if food['crude_protein'] is None:
        food['crude_protein'] = 0
    if food['crude_fat'] is None:
        food['crude_fat'] = 0

    food.update((k, v) for k, v in food.items() if v is not None)  # remove none key value pairs
    food_model, created = FoodModel.objects.get_or_create(**food)
    if created:
        known_as_names = ws.cell(row=i, column=4).value.split(':')
        for name in known_as_names:
            print(name)
            FoodNameModel.objects.create(known_as_name=name, food=food_model)  # type: FoodNameModel

            # Event model
            event = EventModel()
            event.phrase = name
            event.callback = 'ConsultFood'
            event.action = 'READ'
            event.save()
