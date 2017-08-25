import openpyxl
from openpyxl.worksheet import Worksheet

from consult_food.models import TFDAModel, NutritionTuple, NutritionModel, TaiwanSnackModel, ICookIngredientModel
from consult_food.scripts import create_snacks_to_model, create_food_images, create_icook_ingredients


def create_fda_food_name_model():
    wb = openpyxl.load_workbook("../chat_table/consult_food_database.xlsx", data_only=True)
    ws = wb.active  # type: Worksheet
    food = {
        'integration_number': '',
        'food_type': '',
        'sample_name': '',
        'modified_calorie': '',
        'crude_protein': '',
        'crude_fat': '',
        'carbohydrates': '',
        'dietary_fiber': '',
        'metabolic_carbohydrates': ''
    }

    for i in range(2, ws.max_row + 1):
        print(ws.cell(row=i, column=1).value)
        food['integration_number'] = ws.cell(row=i, column=1).value
        food['food_type'] = ws.cell(row=i, column=2).value
        food['sample_name'] = ws.cell(row=i, column=3).value
        food['modified_calorie'] = ws.cell(row=i, column=5).value
        food['crude_protein'] = ws.cell(row=i, column=6).value
        food['crude_fat'] = ws.cell(row=i, column=7).value
        food['carbohydrates'] = ws.cell(row=i, column=8).value
        food['dietary_fiber'] = ws.cell(row=i, column=9).value
        food['metabolic_carbohydrates'] = ws.cell(row=i, column=10).value

        if food['crude_protein'] is None:
            food['crude_protein'] = 0
        if food['crude_fat'] is None:
            food['crude_fat'] = 0

        food.update((k, v) for k, v in food.items() if v is not None)  # remove none key value pairs
        food_model = TFDAModel.objects.create(**food)
        known_as_names = ws.cell(row=i, column=4).value.split(':')
        known_as_names = [s.rstrip().lstrip() for s in known_as_names]
        food_model.synonyms.create(synonym=food_model.sample_name)
        for name in known_as_names:
            food_model.synonyms.create(synonym=name)

        nutrition = get_nutrition(food_model)
        food_model.nutrition = NutritionModel.objects.create(**nutrition._asdict())
        food_model.save()


def get_nutrition(food_model: TFDAModel) -> NutritionTuple:
    protein = food_model.crude_protein
    fat = food_model.crude_fat
    carbohydrates = food_model.carbohydrates

    fruit_amount: float = 0.0
    vegetable_amount: float = 0.0
    grain_amount: float = 0.0
    protein_food_amount: float = 0.0
    diary_amount: float = 0.0
    oil_amount: float = 0.0

    if food_model.food_type in ['穀物類', '澱粉類']:
        grain_amount = food_model.modified_calorie / 70.0

    elif food_model.food_type in ['肉類', '蛋類', '魚貝類']:
        protein_food_amount = food_model.modified_calorie / 75.0

    elif food_model.food_type == '乳品類':
        diary_amount = food_model.modified_calorie / 120.0

    elif food_model.food_type in ['菇類', '蔬菜類', '藻類']:
        vegetable_amount = food_model.modified_calorie / 25.0

    elif food_model.food_type == '水果類':
        fruit_amount = food_model.modified_calorie / 60.0

    elif food_model.food_type in ['油脂類', '堅果及種子類']:
        oil_amount = food_model.modified_calorie / 45.0

    return NutritionTuple(
        name=food_model.sample_name,
        gram=100,
        calories=calc_calories(protein, fat, carbohydrates),
        protein=protein,
        fat=fat,
        carbohydrates=carbohydrates,
        grain_amount=grain_amount,
        protein_food_amount=protein_food_amount,
        diary_amount=diary_amount,
        vegetable_amount=vegetable_amount,
        fruit_amount=fruit_amount,
        oil_amount=oil_amount
    )


def calc_calories(protein: float, fat: float, carbohydrates: float):
    return protein * 4.0 + fat * 9.0 + carbohydrates * 4.0


def delete_all_models():
    TFDAModel.objects.all().delete()
    TaiwanSnackModel.objects.all().delete()
    ICookIngredientModel.objects.all().delete()
    NutritionModel.objects.all().delete()


def run():
    delete_all_models()
    create_fda_food_name_model()
    create_snacks_to_model.run()
    create_food_images.run()
    create_icook_ingredients.run()
