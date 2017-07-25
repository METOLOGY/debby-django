import openpyxl
from openpyxl.worksheet import Worksheet

from consult_food.models import FoodNameModel, FoodModel


def run():
    wb = openpyxl.load_workbook('../chat_table/consult_food_database.xlsx')
    ws: Worksheet = wb.active

    food_list = list()
    for i in range(2, ws.max_row + 1):
        sample_name = ws.cell(row=i, column=3).value
        synonyms_raw = ws.cell(row=i, column=4).value
        synonyms = synonyms_raw.split(':')
        synonyms = [s.rstrip().lstrip() for s in synonyms]
        food_list.append({'sample_name': sample_name,
                          'synonyms': synonyms})

    for food in food_list:
        print('food: {}'.format(food['sample_name']))
        food_queries = FoodModel.objects.filter(sample_name=food['sample_name'])
        for synonym in food['synonyms']:
            print(synonym)
            food_name_queries = FoodNameModel.objects.filter(known_as_name=synonym)
            if not food_name_queries:
                FoodNameModel.objects.create(known_as_name=synonym, food=food_queries[0])
