from typing import NamedTuple

import openpyxl

from consult_food.models import ICookIngredientModel, NutritionModel, NutritionTuple


class ICookOthers(NamedTuple):
    name: str
    source: str
    gram: float = None
    calories: float = None
    protein: float = None
    fat: float = None
    carbohydrates: float = None
    sodium: float = None


def decode_is_empty(value):
    if value == "-":
        return None
    else:
        return value


def run():
    wb = openpyxl.load_workbook("../chat_table/icook 食物雜項表.xlsx", data_only=True)
    ws = wb.get_sheet_by_name("需要登打的表")

    raw_list = []
    row = 2
    while ws.cell(row=row, column=2).value:
        name = ws.cell(row=row, column=2).value
        calories = ws.cell(row=row, column=4).value
        protein = ws.cell(row=row, column=5).value
        fat = ws.cell(row=row, column=6).value
        carbohydrates = ws.cell(row=row, column=7).value
        sodium = ws.cell(row=row, column=8).value
        source = ws.cell(row=row, column=9).value
        if not source:
            source = 'Google'
        elif not source.startswith('https'):
            source = 'Google'

        others = ICookOthers(name=name,
                             source=source,
                             gram=100,
                             calories=decode_is_empty(calories),
                             protein=decode_is_empty(protein),
                             fat=decode_is_empty(fat),
                             carbohydrates=decode_is_empty(carbohydrates),
                             sodium=decode_is_empty(sodium)
                             )

        raw_list.append(others)
        row += 1

    for other in raw_list:  # type: ICookOthers
        print(other.name)
        queries = ICookIngredientModel.objects.filter(name=other.name)
        if queries:
            print('SKIP')
            continue
        else:
            print('Create')
            nutrition = NutritionTuple(name=other.name,
                                       gram=other.gram,
                                       calories=other.calories,
                                       protein=other.protein,
                                       fat=other.fat,
                                       carbohydrates=other.carbohydrates)
            if nutrition.functions().is_valid():
                nutrition_model = NutritionModel.objects.create(**nutrition._asdict())
                nutrition_model.make_calories_image()
                nutrition_model.save()
            else:
                nutrition_model = None
            ICookIngredientModel.objects.create(name=other.name,
                                                nutrition=nutrition_model,
                                                source=other.source,
                                                gram=other.gram,
                                                calories=other.calories,
                                                protein=other.protein,
                                                fat=other.fat,
                                                carbohydrates=other.carbohydrates,
                                                sodium=other.sodium)
