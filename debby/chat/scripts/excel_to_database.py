from openpyxl import load_workbook
from chat.models import ChatModel
from line.models import EventModel

wb2 = load_workbook('./chat/scripts/chat_database_v1.xlsx')  # path to .xlsx
chat_table = wb2['問候']  # sheet name

for sheet_name in ['問候', '字母', '情緒']:
    chat_table = wb2[sheet_name]  # sheet name

    for num, line in enumerate(chat_table.iter_rows()):
        if line[0].value != None and num > 0:
            question = str(line[0].value)
            for num, phrase in enumerate(line):
                if phrase.value != None:
                    if num > 0:
                        answer = phrase.value

                        # Chat model
                        chat = ChatModel()
                        chat.phrase = question
                        chat.answer = answer
                        chat.save()

            # Event model
            event = EventModel()
            event.phrase = question
            event.callback = 'Chat'
            event.action = 'READ'
            event.save()
