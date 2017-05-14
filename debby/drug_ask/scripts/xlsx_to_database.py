import openpyxl
from openpyxl.worksheet.read_only import ReadOnlyWorksheet

from drug_ask.models import DrugTypeModel, DrugDetailModel
from line.models import EventModel

wb = openpyxl.load_workbook("./chat_table/drug_ask_database.xlsx", data_only=True)
ws = wb["Drugs"]  # type: ReadOnlyWorksheet
drug = {
    "question": "",
    "user_choice": "",
    "answer": "",
    "type": "",
}

for i in range(2, ws.max_row):
    if not ws.cell(row=i, column=1).value:
        break
    print(i)
    drug["question"] = ws.cell(row=i, column=1).value
    user_choice = ws.cell(row=i, column=2).value
    if user_choice:
        user_choice = user_choice[2:]  # 1.Acarbose志樂恆 -> Acarbose志樂恆
    drug["user_choice"] = user_choice
    drug["answer"] = ws.cell(row=i, column=3).value
    drug["type"] = ws.cell(row=i, column=4).value
    drug.update((k, v) for k, v in drug.items() if v is not None)  # remove none key value pairs
    DrugTypeModel.objects.get_or_create(**drug)

    # Event model
    event = EventModel()
    event.phrase = drug["question"]
    event.callback = "DrugAsk"
    event.action = "READ"
    event.save()

event = EventModel()
event.phrase = "藥物資訊查詢"
event.callback = "DrugAsk"
event.action = "READ_FROM_MENU"
event.save()

ws = wb["Inquiry"]  # type: ReadOnlyWorksheet
detail = {
    "type": "",
    "mechanism": "",
    "side_effect": "",
    "taboo": "",
    "awareness": ""
}
for i in range(2, ws.max_column + 1):
    if not ws.cell(row=2, column=i).value:
        break
    print(i)
    detail["type"] = ws.cell(row=1, column=i).value
    detail["mechanism"] = ws.cell(row=2, column=i).value
    detail["side_effect"] = ws.cell(row=3, column=i).value
    detail["taboo"] = ws.cell(row=4, column=i).value
    detail["awareness"] = ws.cell(row=5, column=i).value
    detail.update((k, v) for k, v in detail.items() if v is not None)  # remove none key value pairs
    DrugDetailModel.objects.get_or_create(**detail)
